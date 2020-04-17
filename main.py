from sklearn.neural_network import MLPRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.linear_model import LinearRegression
from sklearn.metrics import accuracy_score, mean_squared_error, r2_score
from sklearn.externals import joblib
import numpy as np
import openpyxl
import math
from tkinter import Tk, filedialog, messagebox, Frame, Menu, StringVar, Label, Button, Entry


class App:
    def __init__(self, master):
        self.material_list = []
        self.result_list = []
        self.input_str = []
        self.output_str = []

        self.frame = Frame(master)
        self.frame.grid(row=1)

        menu_bar = Menu(master)

        menu_load_material = Menu(menu_bar, tearoff=0)
        menu_load_material.add_command(label='导入材料名称', command=self.load_material)
        menu_bar.add_cascade(label='导入材料名称', menu=menu_load_material)

        menu_train = Menu(menu_bar, tearoff=0)
        menu_train.add_command(label='神经网络(Neural Network)', command=self.neural_network_regression)
        menu_train.add_command(label='决策树回归(Decision Tree Regression)', command=self.decision_tree_regression)
        menu_train.add_command(label='线性回归(Linear Regression)', command=self.linear_regression)
        # menu_train.add_command(label='支持向量回归(Support Vector Regression)', command=self.support_vector_regression)
        # menu_train.add_command(label='K近邻回归(K Neighbors Regression)', command=self.k_neighbors_regression)
        # menu_train.add_command(label='随机森林回归(Random Forest Regression)', command=self.random_forest_regression)
        # menu_train.add_command(label='AdaBoost回归(AdaBoost Regression)', command=self.adaboost_regression)
        menu_bar.add_cascade(label='输入数据/训练模型', menu=menu_train)

        self.show_model_exist = StringVar()
        Label(master, textvariable=self.show_model_exist).grid(row=0, columnspan=4)
        try:
            self.model = joblib.load('model.m')
            self.show_model_exist.set('模型已加载！可以输入数据进行预测！')
        except:
            self.show_model_exist.set('模型不存在！请先输入训练数据训练模型！')

        # self.review_button = Button(master, text='查看训练数据', command=self.review_data)
        #         self.review_button.grid(row=27, column=1, columnspan=2)

        self.test_button = Button(master, text='添加一行数据', command=self.test)
        self.test_button.grid(row=27, column=3, columnspan=2)

        self.submit_button = Button(master, text='预测', command=self.predict)
        self.submit_button.grid(row=27, column=5, columnspan=2)

        master.config(menu=menu_bar)

    def load_material(self):
        for widget in self.frame.winfo_children():
            widget.destroy()
        self.material_list = []
        self.result_list = []
        file_name = filedialog.askopenfilename(filetypes=[('XLSX', '.xlsx'), ('XLS', '.xls')])
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Sheet1']
        for material_name in sheet[1]:
            self.material_list.append(material_name.value)
        for result_name in sheet[2]:
            if result_name.value == None:
                break
            self.result_list.append(result_name.value)

        for index, item in enumerate(self.material_list):
            Label(self.frame, text=item).grid(row=index % 18 + 1, column=math.floor(index / 18) * 2)
            string_var = StringVar()
            string_var.set(0)
            entry = Entry(self.frame, textvariable=string_var, width=8)
            # entry.bind('<Key>', self.input_change)
            entry.bind('<FocusIn>', self.input_change)
            entry.grid(row=index % 18 + 1, column=math.floor(index / 18) * 2 + 1)
            self.input_str.append(string_var)
        Label(self.frame, text='总和').grid(row=18, column=math.floor(len(self.material_list) / 18) * 2)
        self.sum_str = StringVar()
        self.sum_str.set(0)
        Label(self.frame, textvariable=self.sum_str).grid(row=18, column=math.floor(len(self.material_list) / 18) * 2 + 1)

        Label(self.frame, text='结果：').grid(row=1, column=math.ceil(len(self.material_list) / 18) * 2)
        for index, item in enumerate(self.result_list):
            Label(self.frame, text=item).grid(row=index + 1, column=7)
            string_var = StringVar()
            string_var.set(0)
            entry = Label(self.frame, textvariable=string_var, width=6)
            entry.grid(row=index + 1, column=8)
            self.output_str.append(string_var)

    def load_data(self):
        n_material = len(self.material_list)
        n_result = len(self.result_list)
        material_index = np.zeros(n_material)
        result_index = np.zeros(n_result)
        file_name = filedialog.askopenfilename(filetypes=[('XLSX', '.xlsx'), ('XLS', '.xls')])
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Sheet1']
        nrows = sheet.max_row
        ncols = sheet.max_column
        nSamples = nrows - 1
        for i in range(ncols):
            if sheet.cell(1, i + 1).value in self.material_list:
                index = np.where(np.array(self.material_list) == sheet.cell(1, i + 1).value)[0][0]
                material_index[index] = i + 1
            if sheet.cell(1, i + 1).value in self.result_list:
                index = np.where(np.array(self.result_list) == sheet.cell(1, i + 1).value)[0][0]
                result_index[index] = i + 1
        data_x = np.zeros(shape=(nSamples, n_material), dtype=np.float64)
        data_y = np.zeros(shape=(nSamples, n_result), dtype=np.float64)
        for i in range(nSamples):
            for j in range(n_material):
                data_x[i, j] = sheet.cell(i + 2, material_index[j]).value
            for j in range(n_result):
                data_y[i, j] = sheet.cell(i + 2, result_index[j]).value
        return data_x, data_y

    def neural_network_regression(self):
        data_x, data_y = self.load_data()
        model = MLPRegressor(activation='relu', hidden_layer_sizes=(20, 10), max_iter=1000)
        model.fit(data_x, data_y)
        joblib.dump(model, 'model.m')
        messagebox.showinfo('成功！', '神经网络模型训练完成！')
        self.show_model_exist.set('模型已加载！可以输入数据进行预测！')

    def decision_tree_regression(self):
        data_x, data_y = self.load_data()
        model = DecisionTreeRegressor()
        model.fit(data_x, data_y)
        joblib.dump(model, 'model.m')
        messagebox.showinfo('成功！', '决策树模型训练完成！')
        self.show_model_exist.set('模型已加载！可以输入数据进行预测！')

    def linear_regression(self):
        data_x, data_y = self.load_data()
        model = LinearRegression()
        model.fit(data_x, data_y)
        joblib.dump(model, 'model.m')
        messagebox.showinfo('成功！', '线性回归模型训练完成！')
        self.show_model_exist.set('模型已加载！可以输入数据进行预测！')
        

    def predict(self):
        try:
            self.model = joblib.load('model.m')
        except:
            self.show_model_exist.set('模型不存在！请先输入训练数据训练模型！')
            messagebox.showerror('错误！', '无法加载模型，请先训练模型！')
            return

        n_material = len(self.material_list)
        pred_x = np.zeros(shape=(1, n_material), dtype=np.float64)
        for i in range(n_material):
            pred_x[0, i] = float(self.input_str[i].get())
        try:
            pred_y = self.model.predict(pred_x)
            for i in range(11):
                self.output_str[i].set('%.3f' % pred_y[0, i])
        except:
            messagebox.showerror('错误！', '预测数据与模型输入不符，请重新训练再预测！')

    def input_change(self, event):
        sum_num = 0
        for entry in self.input_str:
            if type(eval(entry.get())) == int or type(eval(entry.get())) == float:
                sum_num += float(entry.get())
            else:
                self.sum_str.set('NaN')
                return
        self.sum_str.set(sum_num)

    def test(self):
        n_material = len(self.material_list)
        n_result = len(self.result_list)
        material_index = np.zeros(n_material)
        file_name = filedialog.askopenfilename(filetypes=[('XLSX', '.xlsx'), ('XLS', '.xls')])
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook['Sheet1']
        ncols = sheet.max_column
        for i in range(ncols):
            if sheet.cell(1, i + 1).value in self.material_list:
                index = np.where(np.array(self.material_list) == sheet.cell(1, i + 1).value)[0][0]
                material_index[index] = i + 1
        for i in range(n_material):
            if material_index[i] == 0:
                self.input_str[i].set(0)
            else:
                self.input_str[i].set(float(sheet.cell(2, material_index[i]).value))


if __name__ == '__main__':
    master = Tk()
    master.title('材料混合试验')
    app = App(master)
    master.mainloop()